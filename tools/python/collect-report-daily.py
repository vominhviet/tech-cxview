import discord
import asyncio
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
# Định nghĩa token của bot
TOKEN = '1234555555'

# Định nghĩa ID server (guild)
GUILD_ID = 123455555  # Thay bằng ID server của bạn

# Định nghĩa channel ID muốn kiểm tra
CHANNEL_ID = 12345555  # Thay bằng ID channel của bạn

# Thiết lập client
intents = discord.Intents.default()
intents.message_content = True  # Kích hoạt quyền truy cập nội dung tin nhắn

client = discord.Client(intents=intents)

# Hàm tìm thread theo ngày trong channel cụ thể
async def get_thread_by_date(channel, date_str):
    print(f"Kiểm tra channel: {channel.name} (ID: {channel.id})")
    # Duyệt qua các thread trong channel
    for thread in channel.threads:
        if date_str in thread.name:
            return thread
    return None  # Nếu không tìm thấy thread phù hợp

# Hàm lọc tin nhắn
def contains_keywords(message_content):
    keywords = ['hôm nay', 'hôm trước', 'hôm qua', 'khó khăn']
    return any(keyword in message_content.lower() for keyword in keywords)

@client.event
async def on_ready():
    print(f'Logged in as {client.user}')
    
    # Lấy guild (server) từ ID
    guild = client.get_guild(GUILD_ID)
    if guild is None:
        print(f"Không tìm thấy server với ID {GUILD_ID}")
        return

    # Tìm channel theo ID đã định nghĩa sẵn
    channel = guild.get_channel(CHANNEL_ID)
    if channel is None:
        print(f"Không tìm thấy kênh với ID {CHANNEL_ID}")
        return

    # Lấy ngày hôm nay và chuyển thành chuỗi theo định dạng 'DD-MM-YYYY'
    today_date = datetime.utcnow().date()
    date_str = today_date.strftime('%d-%m-%Y')  # Chuyển thành định dạng 'DD-MM-YYYY'

    print(f"Tìm kiếm thread với tên: {date_str}")

    # Tìm thread trong kênh theo ngày hôm nay
    thread = await get_thread_by_date(channel, date_str)
    if thread is None:
        print(f"Không tìm thấy thread cho ngày {date_str}")
        return

    # Lấy các tin nhắn từ thread
    filtered_messages = []

    # Lấy các tin nhắn từ thread
    async for message in thread.history(limit=1000):  # Điều chỉnh giới hạn nếu cần
        # Kiểm tra nếu tin nhắn có từ khóa
        if contains_keywords(message.content):
            # Lưu thông tin tin nhắn vào danh sách
            message_data = {
                'Tên': message.author.display_name,
                'Ngày': message.created_at.strftime('%Y-%m-%d'),
                'Nội dung báo cáo': message.content
            }

            filtered_messages.append(message_data)

    # Chuyển danh sách thành DataFrame
    df = pd.DataFrame(filtered_messages)

    first_message_date = df.iloc[0]['Ngày'] if not df.empty else 'no_data'
    filename = f'report_daily_{first_message_date}.xlsx'  # Tạo tên file từ ngày của tin nhắn đầu tiên
    sheet_name = f'daily_{first_message_date}'
    
    # Lưu vào file Excel
    if not df.empty:
        df.to_excel(filename, index=False, engine='openpyxl', sheet_name=sheet_name)

        # Mở tệp Excel vừa tạo
        wb = load_workbook(filename)
        sheet = wb[sheet_name]

        # Thêm kiểu font cho tiêu đề
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Chỉnh sửa kích thước cột
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # lấy chữ cái của cột
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Thêm một chút không gian
            sheet.column_dimensions[column].width = adjusted_width

        # Lưu lại file Excel đã chỉnh sửa
        wb.save(filename)
        print(f"Đã lưu tin nhắn vào file '{filename}' và đã định dạng lại.")
    else:
        print("Không có tin nhắn nào thỏa mãn điều kiện.")
    # # Gửi file Excel vào thread
    # with open(filename, 'rb') as file:
    #     await thread.send("Tổng hợp báo cáo daily:", file=discord.File(file, filename=filename))
    #     print(f"Đã gửi file '{filename}' vào thread.")
# Chạy bot
client.run(TOKEN)
